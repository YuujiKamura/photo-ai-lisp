#!/usr/bin/env python3
"""Probe: does openpyxl embed JPEGs losslessly into an XLSX?

Usage:
    python demo/xlsx-lossless-probe.py [DIR|--synth]

Writes `demo/xlsx-probe-out.xlsx`. Exit 0 if every original JPEG's
sha256 appears verbatim in `xl/media/` of the saved workbook, else 1.

Context: issue #16 (case-container format). Verifying the hypothesis
that a single `.xlsx` per case can hold originals losslessly.
"""

from __future__ import annotations

import hashlib
import os
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

MAX_PHOTOS = 10


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_of_bytes(buf: bytes) -> str:
    return hashlib.sha256(buf).hexdigest()


def collect_jpegs(dir_path: Path) -> List[Path]:
    if not dir_path.is_dir():
        return []
    out: List[Path] = []
    for p in sorted(dir_path.rglob("*")):
        if p.is_file() and p.suffix.lower() in (".jpg", ".jpeg"):
            out.append(p)
            if len(out) >= MAX_PHOTOS:
                break
    return out


def synthesize_jpegs(dst_dir: Path, count: int = 5) -> List[Path]:
    from PIL import Image, ImageDraw
    dst_dir.mkdir(parents=True, exist_ok=True)
    paths: List[Path] = []
    for i in range(count):
        im = Image.new("RGB", (640, 480), color=(20 + i * 40, 90, 200 - i * 30))
        d = ImageDraw.Draw(im)
        d.rectangle((20, 20, 620, 460), outline=(255, 255, 255), width=3)
        d.text((40, 40), f"synth jpeg #{i} photo-ai-lisp issue-16 probe",
               fill=(255, 255, 255))
        out = dst_dir / f"synth-{i:02d}.jpg"
        im.save(out, format="JPEG", quality=88)
        paths.append(out)
    return paths


def pick_source() -> tuple[List[Path], str]:
    home = Path(os.environ.get("USERPROFILE") or os.path.expanduser("~"))

    if len(sys.argv) > 1 and sys.argv[1] != "--synth":
        arg = Path(sys.argv[1])
        photos = collect_jpegs(arg)
        if photos:
            return photos, f"arg:{arg}"

    if len(sys.argv) <= 1:
        for candidate in (home / "apr2026-snapshot",
                          home / "Pictures",
                          home / "Downloads"):
            photos = collect_jpegs(candidate)
            if photos:
                return photos, f"auto:{candidate}"

    tmp = Path(tempfile.mkdtemp(prefix="xlsx-probe-synth-"))
    return synthesize_jpegs(tmp, 5), f"synth:{tmp}"


def build_workbook(photos: List[Path], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Photos"
    ws.append(["basename", "sha256", "bytes"])
    for i, p in enumerate(photos):
        ws.cell(row=i + 2, column=1, value=p.name)
        ws.cell(row=i + 2, column=2, value=sha256_of(p))
        ws.cell(row=i + 2, column=3, value=p.stat().st_size)
    anchor_row = 2 + len(photos) + 2
    for i, p in enumerate(photos):
        img = XLImage(str(p))
        anchor = f"{get_column_letter(1)}{anchor_row + i * 30}"
        ws.add_image(img, anchor)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def extract_media_hashes(xlsx_path: Path) -> List[tuple[str, str, int]]:
    out: List[tuple[str, str, int]] = []
    with zipfile.ZipFile(xlsx_path) as z:
        for name in z.namelist():
            if name.startswith("xl/media/"):
                data = z.read(name)
                out.append((name, sha256_of_bytes(data), len(data)))
    return out


def main() -> int:
    photos, source_tag = pick_source()
    if not photos:
        print("FAIL: no JPEGs available, and synth fallback did not produce any",
              file=sys.stderr)
        return 1

    print(f"source: {source_tag}")
    print(f"photos: {len(photos)}")

    originals = []
    total_src = 0
    for p in photos:
        h = sha256_of(p)
        sz = p.stat().st_size
        total_src += sz
        originals.append((p.name, h, sz))
        print(f"  src  {p.name:40s} {h[:16]}  {sz:>10d} bytes")

    out_xlsx = Path(__file__).resolve().parent / "xlsx-probe-out.xlsx"
    build_workbook(photos, out_xlsx)
    xlsx_size = out_xlsx.stat().st_size
    print(f"wrote: {out_xlsx}  ({xlsx_size} bytes)")

    media = extract_media_hashes(out_xlsx)
    print(f"xl/media entries: {len(media)}")
    for name, h, sz in media:
        print(f"  xlsx {name:40s} {h[:16]}  {sz:>10d} bytes")

    orig_hashes = {h for _, h, _ in originals}
    media_hashes = {h for _, h, _ in media}

    passes = 0
    fails: List[str] = []
    for name, h, _ in originals:
        if h in media_hashes:
            print(f"PASS {name}  {h[:16]}")
            passes += 1
        else:
            print(f"FAIL {name}  {h[:16]} not found in xl/media/")
            fails.append(name)

    unexpected = media_hashes - orig_hashes
    if unexpected:
        print(f"WARN unexpected media hashes (not from originals): {len(unexpected)}")
        for u in unexpected:
            print(f"  extra {u[:16]}")

    overhead = (xlsx_size - total_src) / max(total_src, 1)
    print(f"total originals: {total_src} bytes")
    print(f"total xlsx:      {xlsx_size} bytes")
    print(f"overhead ratio:  {overhead:+.4f}  "
          f"({'xlsx larger' if xlsx_size > total_src else 'xlsx smaller'})")
    print(f"result: {passes}/{len(originals)} PASS")

    return 0 if passes == len(originals) and not fails else 1


if __name__ == "__main__":
    raise SystemExit(main())
